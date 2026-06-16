"""
import_shifts.py
Reads Shift_Plan Excel, parses shift values, upserts into ShiftEntries.
Usage: python import_shifts.py [path_to_xlsx]
"""

import sys
import re
import traceback
from datetime import date, timedelta

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run:  pip install openpyxl")
    sys.exit(1)

try:
    import pyodbc
except ImportError:
    print("ERROR: pyodbc not installed. Run:  pip install pyodbc")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_FILE  = sys.argv[1] if len(sys.argv) > 1 else r"C:\GSDDashboard\Shift_Plan_2026__9_.xlsx"
CUTOFF      = date(2026, 6, 17)          # skip dates before this
SOURCE_SHEET = "EXCEL"
SOURCE_MOD   = "EXCEL_IMPORT"

CONN_STR = (
    "DRIVER={SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=GSDDashboard;"
    "Trusted_Connection=yes;"
)

# ---------------------------------------------------------------------------
# Excel date conversion
# ---------------------------------------------------------------------------
_EXCEL_EPOCH = date(1899, 12, 30)

def excel_serial_to_date(val):
    """Convert Excel date serial or datetime to date. Returns None on failure."""
    if val is None:
        return None
    from datetime import datetime as _dt
    if isinstance(val, _dt):        # check datetime FIRST — it subclasses date
        return val.date()
    if isinstance(val, date):
        return val
    try:
        n = int(val)
        if n < 1:
            return None
        return _EXCEL_EPOCH + timedelta(days=n)
    except Exception:
        return None

# ---------------------------------------------------------------------------
# Shift value parser
# ---------------------------------------------------------------------------
_TIME_RE = re.compile(r'(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})')

def parse_shift(raw):
    """
    Returns (shift_type, shift_start, shift_end) or None to skip.
    shift_start / shift_end may be None for absence types.
    """
    if raw is None:
        return None
    if isinstance(raw, date):
        return None                          # date cell in data area – skip

    s = str(raw).strip()
    if not s or s in ('/', '-', '0', 'None'):
        return None

    su = s.upper()

    # WIC duty  – "WIC 08:00 - 17:00"  or just "WIC"
    if su.startswith('WIC'):
        m = _TIME_RE.search(s)
        if m:
            return ('WIC_DUTY', _fmt_time(m.group(1)), _fmt_time(m.group(2)))
        return ('WIC_DUTY', None, None)

    # Training
    if 'TRAINING' in su:
        return ('TRAINING', None, None)

    # Half-AL variants  "HAL *" / "* HAL" / "HALF_AL"
    if 'HAL' in su:
        return ('HALF_AL', None, None)

    # Half-SL / HSL
    if 'HSL' in su:
        return ('SL', None, None)

    # Half-CD / HCD
    if 'HCD' in su:
        return ('CD', None, None)

    # Straight absence codes
    if su in ('AL', 'AL*', '*AL'):               return ('AL',       None, None)
    if su in ('SL', 'SL*', '*SL', 'SICK'):       return ('SL',       None, None)
    if su in ('UL', 'UL*', '*UL'):               return ('UL',       None, None)
    if su in ('OFF', 'OFFWE', 'OFF_WEEKEND',
              'OFF*', 'OFFDAY'):                 return ('OFF',      None, None)
    if su in ('CD', 'CD*', '*CD'):               return ('CD',       None, None)
    if su in ('PH', 'PH*'):                      return ('PH',       None, None)
    if su in ('LPH',):                           return ('LPH',      None, None)
    if su in ('CO',):                            return ('CO',       None, None)
    if su in ('OL',):                            return ('OL',       None, None)
    if su in ('RESIGNED',):                      return ('RESIGNED', None, None)

    # Pure time range → WORKING  "08:00 - 17:00"
    m = _TIME_RE.search(s)
    if m:
        return ('WORKING', _fmt_time(m.group(1)), _fmt_time(m.group(2)))

    return None   # unrecognised – skip silently

def _fmt_time(t):
    """Ensure time string is zero-padded HH:MM."""
    parts = t.split(':')
    return f"{int(parts[0]):02d}:{parts[1]}"

def parse_date_string(s):
    """Parse a date string from a COM-exported CSV. Handles serial ints and
    common locale formats (German DD.MM.YYYY, US MM/DD/YYYY, ISO YYYY-MM-DD)."""
    from datetime import datetime as _dt
    if not s:
        return None
    s = s.strip()
    if not s:
        return None
    # Pure integer → Excel serial
    if s.isdigit():
        n = int(s)
        if 40000 < n < 60000:   # sane range for ~2009-2064
            return _EXCEL_EPOCH + timedelta(days=n)
        return None
    for fmt in ('%d.%m.%Y', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d',
                '%d.%m.%y', '%m/%d/%y', '%d-%m-%Y'):
        try:
            return _dt.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print()
    print("=" * 60)
    print("  GSD Shift Plan Import")
    print("=" * 60)
    print(f"  File   : {EXCEL_FILE}")
    print(f"  Cutoff : {CUTOFF} (skip earlier dates)")
    print()

    is_csv = EXCEL_FILE.lower().endswith('.csv')

    if is_csv:
        # ── CSV path (COM-exported clean file) ───────────────────────────────
        import csv as _csv
        try:
            with open(EXCEL_FILE, encoding='utf-8-sig', newline='') as f:
                all_rows = list(_csv.reader(f))
        except FileNotFoundError:
            print(f"ERROR: CSV not found: {EXCEL_FILE}")
            sys.exit(1)
        except Exception as ex:
            print(f"ERROR reading CSV: {ex}")
            sys.exit(1)

        print(f"  CSV rows     : {len(all_rows)}")
        if all_rows:
            print(f"  Row 1 [1:10] : {all_rows[0][:10]}")
        if len(all_rows) > 1:
            print(f"  Row 2 [1:10] : {all_rows[1][:10]}")

        header_row  = all_rows[0] if all_rows else []
        date_cols   = {}
        skipped_hdr = []

        for col_1based in range(8, len(header_row) + 1):
            v = header_row[col_1based - 1].strip()
            if not v:
                continue
            d = parse_date_string(v)
            if d is None:
                skipped_hdr.append((col_1based, v))
                continue
            if d >= CUTOFF:
                date_cols[col_1based] = d

        if not date_cols:
            print("ERROR: No date columns found on or after cutoff in CSV header.")
            sample = [(i + 8, header_row[i + 7]) for i in range(10) if i + 7 < len(header_row)]
            print(f"  Header cols 8-17 : {sample}")
            sys.exit(1)

        dates_sorted = sorted(date_cols.values())
        print(f"  Date columns : {len(date_cols)}  ({dates_sorted[0]} to {dates_sorted[-1]})")
        if skipped_hdr:
            print(f"  Non-date header cells : {skipped_hdr[:5]}")

        shifts        = []
        skipped_no_id = 0
        skipped_blank = 0
        unknown_vals  = {}

        for raw_row in all_rows[1:]:
            raw_id = raw_row[1].strip() if len(raw_row) > 1 else ''
            if not raw_id:
                skipped_no_id += 1
                continue
            emp_id = raw_id.split('.')[0]   # strip .0 if Excel exported as float
            if not emp_id or emp_id == '0':
                skipped_no_id += 1
                continue

            for col_1based, shift_date in date_cols.items():
                col_idx = col_1based - 1
                if col_idx >= len(raw_row):
                    continue
                cell_val = raw_row[col_idx].strip() or None
                parsed = parse_shift(cell_val)
                if parsed is None:
                    if cell_val and cell_val not in ('/', '-'):
                        unknown_vals[cell_val] = unknown_vals.get(cell_val, 0) + 1
                    skipped_blank += 1
                    continue
                shifts.append((emp_id, shift_date, parsed[0], parsed[1], parsed[2]))

    else:
        # ── XLSX path ────────────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(
                EXCEL_FILE, data_only=True, keep_vba=True, read_only=True
            )
        except FileNotFoundError:
            print(f"ERROR: File not found: {EXCEL_FILE}")
            print("Copy Shift_Plan_2026__9_.xlsx to C:\\GSDDashboard\\ first.")
            sys.exit(1)
        except Exception as ex:
            print(f"ERROR loading workbook: {ex}")
            sys.exit(1)

        print(f"  Sheet names : {wb.sheetnames}")

        TARGET_SHEETS = ['Shift Plan GSD DE', 'Shift Plan WIC', 'Resigned']
        ws = None
        sheet_used = None
        for candidate in TARGET_SHEETS:
            if candidate not in wb.sheetnames:
                continue
            sheet = wb[candidate]
            probe = next(sheet.iter_rows(min_row=1, max_row=1,
                                         min_col=1, max_col=10,
                                         values_only=True), ())
            has_data = any(v is not None for v in probe)
            print(f"  Sheet '{candidate}': probe row1={probe[:6]}  has_data={has_data}")
            if has_data and ws is None:
                ws = sheet
                sheet_used = candidate

        if ws is None:
            print("ERROR: No readable sheet found. All candidate sheets return None.")
            sys.exit(1)

        print(f"\n  Using sheet: '{sheet_used}'")
        print(f"  Dimensions : max_row={ws.max_row}, max_col={ws.max_column}")

        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        date_cols   = {}
        skipped_hdr = []

        for col_1based, v in enumerate(header, start=1):
            if col_1based < 8:
                continue
            if v is None:
                continue
            d = excel_serial_to_date(v)
            if d is None:
                skipped_hdr.append((col_1based, v))
                continue
            if d >= CUTOFF:
                date_cols[col_1based] = d

        if not date_cols:
            print("ERROR: No date columns found on or after cutoff in header row 1.")
            sample = [(i+1, v) for i, v in enumerate(header[7:17]) if v is not None]
            print(f"  Header cols 8-17 sample : {sample}")
            print(f"  Cutoff                  : {CUTOFF}")
            sys.exit(1)

        dates_sorted = sorted(date_cols.values())
        print(f"  Date columns : {len(date_cols)}  "
              f"({dates_sorted[0]} to {dates_sorted[-1]})")
        if skipped_hdr:
            print(f"  Non-date header cells   : {skipped_hdr[:5]}")

        shifts        = []
        skipped_no_id = 0
        skipped_blank = 0
        unknown_vals  = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_id = row[1] if len(row) > 1 else None
            if raw_id is None:
                skipped_no_id += 1
                continue
            emp_id = str(raw_id).strip().split('.')[0]
            if not emp_id or emp_id == '0':
                skipped_no_id += 1
                continue

            for col_1based, shift_date in date_cols.items():
                col_idx = col_1based - 1
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx]
                parsed = parse_shift(cell_val)
                if parsed is None:
                    if cell_val is not None:
                        sv = str(cell_val).strip()
                        if sv and sv not in ('/', '-'):
                            unknown_vals[sv] = unknown_vals.get(sv, 0) + 1
                    skipped_blank += 1
                    continue
                shifts.append((emp_id, shift_date, parsed[0], parsed[1], parsed[2]))

    print(f"  Rows parsed  : {len(shifts)} shift entries")
    print(f"  Skipped      : {skipped_no_id} (no ID), {skipped_blank} (blank/skip cells)")
    if unknown_vals:
        top = sorted(unknown_vals.items(), key=lambda x: -x[1])[:10]
        print(f"  Unknown values (top 10): {top}")
    print()

    if not shifts:
        print("Nothing to import.")
        return

    # ── Connect and upsert ───────────────────────────────────────
    try:
        conn = pyodbc.connect(CONN_STR)
    except Exception as ex:
        print(f"ERROR connecting to SQL Server: {ex}")
        sys.exit(1)

    conn.autocommit = False
    cursor = conn.cursor()

    inserted = 0
    updated  = 0
    errors   = 0

    for emp_id, shift_date, shift_type, shift_start, shift_end in shifts:
        shift_date_s = shift_date.strftime('%Y-%m-%d')
        try:
            cursor.execute(
                "SELECT Id FROM ShiftEntries WHERE EmployeeId=? AND ShiftDate=?",
                emp_id, shift_date_s
            )
            row = cursor.fetchone()
            if row:
                cursor.execute(
                    "UPDATE ShiftEntries "
                    "SET ShiftType=?, ShiftStart=?, ShiftEnd=?, SourceSheet=?, SourceModule=? "
                    "WHERE Id=?",
                    shift_type, shift_start, shift_end, SOURCE_SHEET, SOURCE_MOD, row[0]
                )
                updated += 1
            else:
                cursor.execute(
                    "INSERT INTO ShiftEntries "
                    "(EmployeeId, ShiftDate, ShiftType, ShiftStart, ShiftEnd, "
                    " SourceSheet, SourceModule, AutoGenerated) "
                    "VALUES (?,?,?,?,?,?,?,0)",
                    emp_id, shift_date_s, shift_type, shift_start, shift_end,
                    SOURCE_SHEET, SOURCE_MOD
                )
                inserted += 1
        except Exception as ex:
            errors += 1
            if errors <= 5:
                print(f"  Row error ({emp_id} / {shift_date}): {ex}")

    try:
        conn.commit()
    except Exception as ex:
        conn.rollback()
        print(f"ERROR during commit: {ex}")
        sys.exit(1)
    finally:
        cursor.close()
        conn.close()

    print(f"  Inserted : {inserted}")
    print(f"  Updated  : {updated}")
    if errors:
        print(f"  Errors   : {errors}  (first 5 shown above)")
    print()
    print("=" * 60)
    print("  Import complete")
    print("=" * 60)

if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        sys.exit(1)
